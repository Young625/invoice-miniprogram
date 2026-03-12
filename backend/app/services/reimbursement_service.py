"""报销包生成服务：Excel汇总表 + 报销单PDF + 发票原件打包"""
import os
import io
import zipfile
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors
from pypdf import PdfWriter, PdfReader

import logging

logger = logging.getLogger(__name__)


class ReimbursementService:
    """报销包生成服务"""

    def __init__(self):
        # 注册中文字体（支持多平台）
        self.font_name = 'SimSun'  # 默认字体名
        self.font_registered = False  # 标记是否成功注册中文字体
        try:
            # 尝试多个字体路径
            font_paths = [
                # Windows 字体路径（优先使用 .ttf 文件，避免 .ttc 文件的兼容性问题）
                "C:/Windows/Fonts/simhei.ttf",  # 黑体 (TTF)
                "C:/Windows/Fonts/msyh.ttf",  # 微软雅黑 (TTF)
                "C:/Windows/Fonts/msyhbd.ttf",  # 微软雅黑粗体 (TTF)
                "C:/Windows/Fonts/simsun.ttc",  # 宋体 (TTC)
                "C:/Windows/Fonts/msyh.ttc",  # 微软雅黑 (TTC)
                # Linux 常见字体路径
                "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
                "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
                "/usr/share/fonts/truetype/arphic/uming.ttc",
                # macOS 字体路径
                "/System/Library/Fonts/PingFang.ttc",
                "/System/Library/Fonts/STHeiti Light.ttc",
            ]

            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                        self.font_name = 'ChineseFont'
                        self.font_registered = True
                        logger.info(f"成功注册中文字体: {font_path}")
                        break
                    except Exception as e:
                        logger.warning(f"注册字体失败 {font_path}: {e}")
                        continue

            if not self.font_registered:
                logger.warning("未找到中文字体，将使用默认字体（中文和特殊符号可能无法显示）")

        except Exception as e:
            logger.error(f"字体注册过程出错: {e}")

    def generate_excel_summary(self, invoices: List[Dict[str, Any]], user_info: Dict[str, Any] = None) -> bytes:
        """
        生成发票汇总表 Excel

        Args:
            invoices: 发票列表（已按 items 排序）
            user_info: 用户信息（包含报销人、OA报销单号等）

        Returns:
            Excel 文件的字节内容
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "发票汇总表"

        # 设置列宽
        ws.column_dimensions['A'].width = 12  # 报销人
        ws.column_dimensions['B'].width = 20  # OA报销单号
        ws.column_dimensions['C'].width = 20  # 发票类型
        ws.column_dimensions['D'].width = 18  # 发票代码
        ws.column_dimensions['E'].width = 20  # 发票号码
        ws.column_dimensions['F'].width = 15  # 开票日期
        ws.column_dimensions['G'].width = 30  # 购买方名称
        ws.column_dimensions['H'].width = 30  # 销售方名称
        ws.column_dimensions['I'].width = 20  # 销售方税号
        ws.column_dimensions['J'].width = 25  # 项目名称
        ws.column_dimensions['K'].width = 15  # 发票金额(不含税)
        ws.column_dimensions['L'].width = 12  # 发票税额
        ws.column_dimensions['M'].width = 10  # 发票税率
        ws.column_dimensions['N'].width = 12  # 合税金额
        ws.column_dimensions['O'].width = 15  # 录入日期

        # 标题样式
        title_font = Font(name='Arial', size=11, bold=True)
        title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        title_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头
        headers = [
            '报销人', 'OA报销单号', '发票类型', '发票代码', '发票号码',
            '开票日期', '购买方名称', '销售方名称', '销售方税号', '项目名称',
            '发票金额(不含税)', '发票税额', '发票税率', '合税金额', '录入日期'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
            cell.border = thin_border

        # 数据行
        total_amount = 0
        total_tax = 0
        total_with_tax = 0

        for row_num, invoice in enumerate(invoices, 2):
            # 报销人
            ws.cell(row=row_num, column=1, value=user_info.get('name', '') if user_info else '').border = thin_border

            # OA报销单号
            ws.cell(row=row_num, column=2, value=user_info.get('oa_number', '') if user_info else '').border = thin_border

            # 发票类型
            ws.cell(row=row_num, column=3, value=invoice.get('invoice_type', '')).border = thin_border

            # 发票代码
            ws.cell(row=row_num, column=4, value=invoice.get('invoice_code', '')).border = thin_border

            # 发票号码
            ws.cell(row=row_num, column=5, value=invoice.get('invoice_number', '')).border = thin_border

            # 开票日期
            ws.cell(row=row_num, column=6, value=invoice.get('invoice_date', '')).border = thin_border

            # 购买方名称
            ws.cell(row=row_num, column=7, value=invoice.get('buyer_name', '')).border = thin_border

            # 销售方名称
            ws.cell(row=row_num, column=8, value=invoice.get('seller_name', '')).border = thin_border

            # 销售方税号
            ws.cell(row=row_num, column=9, value=invoice.get('seller_tax_id', '')).border = thin_border

            # 项目名称（items 列表转换为字符串）
            items = invoice.get('items', [])
            project_name_str = ', '.join(items) if items else ''
            ws.cell(row=row_num, column=10, value=project_name_str).border = thin_border

            # 发票金额(不含税)
            amount = invoice.get('amount', 0) or 0
            ws.cell(row=row_num, column=11, value=amount).border = thin_border
            total_amount += amount

            # 发票税额
            tax_amount = invoice.get('tax_amount', 0) or 0
            ws.cell(row=row_num, column=12, value=tax_amount).border = thin_border
            total_tax += tax_amount

            # 发票税率（写入数值 0.03，格式设为百分比，Excel 显示 3%）
            tax_rate = invoice.get('tax_rate')
            tax_rate_cell = ws.cell(row=row_num, column=13, value=tax_rate / 100 if tax_rate is not None else None)
            tax_rate_cell.number_format = '0%'
            tax_rate_cell.border = thin_border

            # 合税金额（价税合计）
            total_amt = invoice.get('total_amount', 0) or 0
            ws.cell(row=row_num, column=14, value=total_amt).border = thin_border
            total_with_tax += total_amt

            # 录入日期
            created_at = invoice.get('created_at', '')
            if created_at:
                # 如果是 datetime 对象，转换为字符串
                if hasattr(created_at, 'strftime'):
                    created_at = created_at.strftime('%Y-%m-%d')
                elif isinstance(created_at, str):
                    # 如果是字符串，尝试提取日期部分
                    created_at = created_at.split('T')[0] if 'T' in created_at else created_at[:10]
            ws.cell(row=row_num, column=15, value=created_at).border = thin_border

        # 合计行
        summary_row = len(invoices) + 2
        ws.cell(row=summary_row, column=10, value='合计：').font = Font(bold=True)
        ws.cell(row=summary_row, column=11, value=total_amount).font = Font(bold=True, color="FF0000")
        ws.cell(row=summary_row, column=12, value=total_tax).font = Font(bold=True, color="FF0000")
        # 税率列不需要合计
        ws.cell(row=summary_row, column=14, value=total_with_tax).font = Font(bold=True, color="FF0000")

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_reimbursement_pdf(
        self,
        invoices: List[Dict[str, Any]],
        user_info: Dict[str, Any]
    ) -> bytes:
        """
        生成报销单 PDF

        Args:
            invoices: 发票列表
            user_info: 用户信息（报销人、部门、事由等）

        Returns:
            PDF 文件的字节内容
        """
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        try:
            # 使用中文字体
            c.setFont(self.font_name, 12)
        except:
            # 如果中文字体不可用，使用默认字体
            c.setFont('Helvetica', 12)
            logger.warning("使用默认字体，中文可能无法正常显示")

        # 标题
        try:
            c.setFont(self.font_name, 20)
        except:
            c.setFont('Helvetica-Bold', 20)
        c.drawCentredString(width / 2, height - 2 * cm, "费用报销单")

        # 报销信息
        y_position = height - 4 * cm
        try:
            c.setFont(self.font_name, 12)
        except:
            c.setFont('Helvetica', 12)

        # 报销人信息
        reimbursement_no = f"BX{datetime.now().strftime('%Y%m%d%H%M%S')}"

        # 只显示已填写的字段
        name = user_info.get('name')
        department = user_info.get('department')

        if name:
            c.drawString(3 * cm, y_position, f"报销人: {name}")
            if department:
                c.drawString(10 * cm, y_position, f"部门: {department}")
                y_position -= 1 * cm
            else:
                y_position -= 1 * cm
        elif department:
            c.drawString(3 * cm, y_position, f"部门: {department}")
            y_position -= 1 * cm

        c.drawString(3 * cm, y_position, f"日期: {datetime.now().strftime('%Y-%m-%d')}")
        c.drawString(10 * cm, y_position, f"单号: {reimbursement_no}")

        # 报销明细表
        y_position -= 2 * cm
        try:
            c.setFont(self.font_name, 12)
        except:
            c.setFont('Helvetica-Bold', 12)
        c.drawString(3 * cm, y_position, "报销明细:")

        # 表头
        y_position -= 0.8 * cm
        try:
            c.setFont(self.font_name, 10)
        except:
            c.setFont('Helvetica', 10)
        c.drawString(3 * cm, y_position, "序号")
        c.drawString(5 * cm, y_position, "发票号码")
        c.drawString(10 * cm, y_position, "金额（元）")
        c.drawString(14 * cm, y_position, "日期")

        # 画线
        c.line(2.5 * cm, y_position - 0.2 * cm, width - 2.5 * cm, y_position - 0.2 * cm)

        # 明细数据
        total_amount = 0
        for idx, invoice in enumerate(invoices, 1):
            y_position -= 0.8 * cm
            if y_position < 5 * cm:  # 换页
                c.showPage()
                y_position = height - 3 * cm
                try:
                    c.setFont(self.font_name, 10)
                except:
                    c.setFont('Helvetica', 10)

            c.drawString(3 * cm, y_position, str(idx))
            c.drawString(5 * cm, y_position, invoice.get('invoice_number', '')[:15])
            amount = invoice.get('total_amount', 0) or 0
            c.drawString(10 * cm, y_position, f"{amount:.2f}")
            c.drawString(14 * cm, y_position, invoice.get('invoice_date', ''))
            total_amount += amount

        # 合计
        y_position -= 1 * cm
        c.line(2.5 * cm, y_position, width - 2.5 * cm, y_position)
        y_position -= 0.8 * cm
        try:
            c.setFont(self.font_name, 12)
        except:
            c.setFont('Helvetica-Bold', 12)
        c.drawString(8 * cm, y_position, f"合计金额: ¥{total_amount:.2f}")

        # 报销事由（仅在填写时显示）
        reason = user_info.get('reason')
        if reason:
            y_position -= 2 * cm
            try:
                c.setFont(self.font_name, 12)
            except:
                c.setFont('Helvetica', 12)
            c.drawString(3 * cm, y_position, f"报销事由: {reason}")

        # 签字栏
        y_position -= 3 * cm
        c.drawString(3 * cm, y_position, "报销人签字: _____________")
        c.drawString(10 * cm, y_position, "日期: _____________")

        y_position -= 1 * cm
        c.drawString(3 * cm, y_position, "部门主管: _____________")
        c.drawString(10 * cm, y_position, "日期: _____________")

        y_position -= 1 * cm
        c.drawString(3 * cm, y_position, "财务审核: _____________")
        c.drawString(10 * cm, y_position, "日期: _____________")

        c.save()
        buffer.seek(0)
        return buffer.getvalue()

    def merge_invoice_pdfs(
        self,
        invoices: List[Dict[str, Any]],
        pdf_dir: str
    ) -> bytes:
        """
        合并所有发票PDF为一个文件

        Args:
            invoices: 发票列表
            pdf_dir: 发票PDF文件存储目录

        Returns:
            合并后的PDF文件字节内容，如果没有可合并的PDF则返回None
        """
        pdf_writer = PdfWriter()
        merged_count = 0

        for idx, invoice in enumerate(invoices, 1):
            pdf_path = invoice.get('pdf_path')
            if pdf_path:
                # 统一路径分隔符
                pdf_path = pdf_path.replace('\\', os.sep).replace('/', os.sep)
                full_pdf_path = os.path.join(pdf_dir, pdf_path)

                if os.path.exists(full_pdf_path):
                    try:
                        # 读取PDF并添加到合并器
                        pdf_reader = PdfReader(full_pdf_path)
                        for page in pdf_reader.pages:
                            pdf_writer.add_page(page)
                        merged_count += 1
                        logger.info(f"已合并发票PDF: {invoice.get('invoice_number')}")
                    except Exception as e:
                        logger.error(f"合并PDF失败: {full_pdf_path}, {e}")
                else:
                    logger.warning(f"发票PDF文件不存在: {full_pdf_path}")

        # 如果有成功合并的PDF，返回字节内容
        if merged_count > 0:
            output = io.BytesIO()
            pdf_writer.write(output)
            output.seek(0)
            logger.info(f"成功合并 {merged_count} 个发票PDF")
            return output.getvalue()
        else:
            logger.warning("没有可合并的发票PDF")
            return None

    def create_reimbursement_package(
        self,
        invoices: List[Dict[str, Any]],
        user_info: Dict[str, Any],
        pdf_dir: str
    ) -> bytes:
        """
        创建完整的报销包（ZIP文件）

        Args:
            invoices: 发票列表
            user_info: 用户信息
            pdf_dir: 发票PDF文件存储目录

        Returns:
            ZIP 文件的字节内容
        """
        # 按 items 排序发票（确保所有文件都使用相同的排序）
        def get_sort_key(invoice):
            items = invoice.get('items', [])
            if not items:
                return 'zzz'  # 空 items 排在最后
            return ', '.join(items)

        invoices = sorted(invoices, key=get_sort_key)

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 1. 添加 Excel 汇总表
            excel_data = self.generate_excel_summary(invoices, user_info)
            zip_file.writestr('发票汇总表.xlsx', excel_data)
            logger.info("已添加 Excel 汇总表")

            # 2. 添加报销单 PDF
            pdf_data = self.generate_reimbursement_pdf(invoices, user_info)
            zip_file.writestr('报销单.pdf', pdf_data)
            logger.info("已添加报销单 PDF")

            # 3. 添加发票原件 PDF
            pdf_count = 0
            for idx, invoice in enumerate(invoices, 1):
                pdf_path = invoice.get('pdf_path')
                if pdf_path:
                    # 将路径分隔符统一转换为当前系统的分隔符
                    # 数据库中可能存储的是Windows风格的路径（使用\）
                    pdf_path = pdf_path.replace('\\', os.sep).replace('/', os.sep)

                    # 拼接完整路径
                    full_pdf_path = os.path.join(pdf_dir, pdf_path)

                    logger.info(f"尝试读取PDF: {full_pdf_path}")
                    full_pdf_path = os.path.join(pdf_dir, pdf_path)

                    if os.path.exists(full_pdf_path):
                        try:
                            with open(full_pdf_path, 'rb') as f:
                                pdf_content = f.read()
                            # 使用序号和发票号码命名
                            invoice_number = invoice.get('invoice_number', 'unknown')
                            filename = f"{idx:02d}_{invoice_number}.pdf"
                            zip_file.writestr(f'发票原件/{filename}', pdf_content)
                            pdf_count += 1
                            logger.info(f"已添加发票原件: {filename}")
                        except Exception as e:
                            logger.error(f"添加发票 PDF 失败: {full_pdf_path}, {e}")
                    else:
                        logger.warning(f"发票 PDF 文件不存在: {full_pdf_path}")
                else:
                    logger.warning(f"发票 {invoice.get('invoice_number')} 没有 PDF 路径")

            logger.info(f"已添加 {pdf_count}/{len(invoices)} 个发票原件 PDF")

            # 4. 添加合并后的发票PDF
            merged_pdf = self.merge_invoice_pdfs(invoices, pdf_dir)
            if merged_pdf:
                zip_file.writestr('发票原件合并.pdf', merged_pdf)
                logger.info("已添加合并后的发票PDF")

        zip_buffer.seek(0)
        return zip_buffer.getvalue()
