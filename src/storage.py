"""存储管理模块：PDF 按日期归档 + Excel 汇总表管理。

PDF 存储格式：data/invoices/YYYY/MM/YYYYMMDD_发票号码_销售方.pdf
Excel 汇总表：data/output/发票汇总.xlsx
"""

import logging
import os
import re
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.utils import get_column_letter

try:
    from .invoice_parser import InvoiceInfo
except ImportError:
    from invoice_parser import InvoiceInfo

logger = logging.getLogger(__name__)

# Excel 列定义
EXCEL_COLUMNS = [
    ("序号", 8),
    ("发票类型", 20),
    ("发票代码", 16),
    ("发票号码", 24),
    ("开票日期", 14),
    ("购买方名称", 30),
    ("购买方税号", 22),
    ("销售方名称", 30),
    ("销售方税号", 22),
    ("商品名称", 40),
    ("金额", 14),
    ("税额", 14),
    ("价税合计", 14),
    ("来源邮件主题", 40),
    ("提取时间", 20),
    ("PDF路径", 50),
]


def _sanitize_filename(name: str) -> str:
    """清理文件名中的非法字符。"""
    # 移除/替换文件名中不允许的字符
    name = re.sub(r'[\\/:*?"<>|\r\n]', "_", name)
    # 限制长度
    return name[:50].strip("_. ")


class StorageManager:
    """存储管理器：PDF 归档和 Excel 汇总。"""

    def __init__(self, invoice_dir: str, excel_path: str):
        self.invoice_dir = invoice_dir
        self.excel_path = excel_path

    def save_pdf(
        self, pdf_data: bytes, info: InvoiceInfo, original_filename: str
    ) -> str:
        """将发票 PDF 保存到按日期组织的目录中。

        存储路径格式：YYYY/MM/YYYYMMDD_发票号码_销售方.pdf

        Args:
            pdf_data: PDF 二进制内容
            info: 解析后的发票信息
            original_filename: 原始文件名（备用）

        Returns:
            保存后的文件路径（相对于 invoice_dir）
        """
        # 确定日期部分
        if info.invoice_date:
            try:
                date = datetime.strptime(info.invoice_date, "%Y-%m-%d")
            except ValueError:
                date = datetime.now()
        else:
            date = datetime.now()

        year_str = date.strftime("%Y")
        month_str = date.strftime("%m")
        date_prefix = date.strftime("%Y%m%d")

        # 构建文件名
        parts = [date_prefix]
        if info.invoice_number:
            parts.append(info.invoice_number)
        if info.seller_name:
            parts.append(_sanitize_filename(info.seller_name))
        else:
            # 使用原始文件名（去掉扩展名）
            base_name = os.path.splitext(original_filename)[0]
            parts.append(_sanitize_filename(base_name))

        filename = "_".join(parts) + ".pdf"

        # 创建目录
        dir_path = os.path.join(self.invoice_dir, year_str, month_str)
        os.makedirs(dir_path, exist_ok=True)

        # 写入文件（如果同名文件已存在则追加序号）
        file_path = os.path.join(dir_path, filename)
        if os.path.exists(file_path):
            base, ext = os.path.splitext(file_path)
            counter = 1
            while os.path.exists(f"{base}_{counter}{ext}"):
                counter += 1
            file_path = f"{base}_{counter}{ext}"

        with open(file_path, "wb") as f:
            f.write(pdf_data)

        # 返回相对路径
        rel_path = os.path.relpath(file_path, os.path.dirname(self.invoice_dir))
        logger.info("PDF 已保存: %s", file_path)
        return rel_path

    def _init_excel(self) -> Workbook:
        """初始化 Excel 工作簿，创建表头。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "发票汇总"

        # 写入表头
        header_font = Font(bold=True, size=11)
        for col_idx, (col_name, col_width) in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # 冻结首行
        ws.freeze_panes = "A2"

        return wb

    def _set_text_format(self, ws, row: int):
        """将发票号码和税号列设为文本格式，防止科学计数法。"""
        # 发票代码(3列)、发票号码(4列)、购买方税号(7列)、销售方税号(9列)
        text_columns = [3, 4, 7, 9]
        for col in text_columns:
            cell = ws.cell(row=row, column=col)
            cell.number_format = numbers.FORMAT_TEXT
            # 确保值为字符串
            if cell.value and isinstance(cell.value, (int, float)):
                cell.value = str(cell.value)

    def append_to_excel(
        self, info: InvoiceInfo, email_subject: str, pdf_rel_path: str
    ):
        """将发票信息追加到 Excel 汇总表。

        Args:
            info: 解析后的发票信息
            email_subject: 来源邮件主题
            pdf_rel_path: PDF 相对路径
        """
        os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)

        # 加载或创建工作簿
        if os.path.exists(self.excel_path):
            wb = load_workbook(self.excel_path)
            ws = wb.active
        else:
            wb = self._init_excel()
            ws = wb.active

        # 计算序号
        next_row = ws.max_row + 1
        seq = next_row - 1  # 减去表头行

        # 商品名称合并
        items_str = "；".join(info.items) if info.items else ""

        # 提取时间
        extract_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 写入数据行
        row_data = [
            seq,
            info.invoice_type,
            info.invoice_code,     # 文本格式
            info.invoice_number,   # 文本格式
            info.invoice_date,
            info.buyer_name,
            info.buyer_tax_id,     # 文本格式
            info.seller_name,
            info.seller_tax_id,    # 文本格式
            items_str,
            info.amount,
            info.tax_amount,
            info.total_amount,
            email_subject,
            extract_time,
            pdf_rel_path,
        ]

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=next_row, column=col_idx, value=value)

        # 设置文本格式列
        self._set_text_format(ws, next_row)

        # 保存
        wb.save(self.excel_path)
        logger.info(
            "Excel 已更新: 第 %d 行, 发票号码 %s", seq, info.invoice_number
        )
